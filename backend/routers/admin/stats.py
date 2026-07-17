# -*- coding: utf-8 -*-
# 管理后台数据统计接口

import datetime
import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.models.admin_user import AdminUser
from backend.schemas.common import ApiResponse
from backend.services.stats_service import stats_service
from backend.utils.admin_auth import (
    deny_observer_export,
    get_current_admin,
    require_role,
)

logger = logging.getLogger(__name__)

router = APIRouter()

_VALID_METRICS = {"new_users", "active_users", "conversation_rounds"}
_VALID_DAYS = {7, 30}
_VALID_REPORT_TYPES = {"user", "conversation", "feature", "ai_performance"}
_REPORT_READ_ROLES = ("super_admin", "ops_admin", "observer")
_LIBLIB_READ_ROLES = ("super_admin", "ai_trainer", "tech_ops", "observer")

# 报表类型 → Excel 表头映射
_REPORT_HEADERS = {
    "user": ["日期", "新增用户数", "活跃用户数"],
    "conversation": ["日期", "对话轮次"],
    "feature": ["日期", "主动消息发送数", "主动消息打开数", "回复率(%)"],
    "ai_performance": ["日期", "人格偏离数", "AI回复数", "偏离率(%)"],
}

# 报表类型 → 每行数据取值字段
_REPORT_FIELDS = {
    "user": ["date", "new_users", "active_users"],
    "conversation": ["date", "conversation_rounds"],
    "feature": ["date", "agent_sent", "agent_opened", "reply_rate"],
    "ai_performance": ["date", "persona_risk_count", "total_count", "deviation_rate"],
}


@router.get("/stats/dashboard")
async def get_dashboard(
    admin_user: AdminUser = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """仪表盘数据（所有角色可访问，按角色过滤返回内容）"""
    data = await stats_service.get_dashboard_data(admin_user.role, db)
    return ApiResponse.ok(data=data)


@router.get(
    "/stats/trend",
    dependencies=[require_role(*_REPORT_READ_ROLES)],
)
async def get_trend(
    metric: str = Query(..., description="指标类型"),
    days: int = Query(..., description="天数"),
    db: AsyncSession = Depends(get_db),
):
    """趋势数据"""
    if metric not in _VALID_METRICS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"metric 参数无效，可选值: {', '.join(_VALID_METRICS)}",
        )
    if days not in _VALID_DAYS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="days 参数无效，可选值: 7, 30",
        )

    data = await stats_service.get_trend_data(metric, days, db)
    return ApiResponse.ok(data=data)


@router.get(
    "/stats/report",
    dependencies=[require_role(*_REPORT_READ_ROLES)],
)
async def get_report(
    report_type: str = Query(..., description="报表类型"),
    start_date: Optional[datetime.date] = Query(None, description="开始日期"),
    end_date: Optional[datetime.date] = Query(None, description="结束日期"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页条数"),
    db: AsyncSession = Depends(get_db),
):
    """报表数据"""
    if report_type not in _VALID_REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"report_type 参数无效，可选值: {', '.join(_VALID_REPORT_TYPES)}",
        )

    today = datetime.date.today()
    if start_date is None:
        start_date = today - datetime.timedelta(days=30)
    if end_date is None:
        end_date = today

    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="结束时间不能早于开始时间",
        )
    if (end_date - start_date).days > 90:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="时间范围不能超过90天，请缩小范围",
        )

    data = await stats_service.get_report_data(
        report_type, start_date, end_date, page, page_size, db
    )
    return ApiResponse.ok(data=data)


@router.post(
    "/stats/report/export",
    dependencies=[
        Depends(deny_observer_export),
        require_role("super_admin", "ops_admin"),
    ],
)
async def export_report(
    report_type: str = Query(..., description="报表类型"),
    start_date: Optional[datetime.date] = Query(None, description="开始日期"),
    end_date: Optional[datetime.date] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
):
    """导出报表为 Excel"""
    if report_type not in _VALID_REPORT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"report_type 参数无效，可选值: {', '.join(_VALID_REPORT_TYPES)}",
        )

    today = datetime.date.today()
    if start_date is None:
        start_date = today - datetime.timedelta(days=30)
    if end_date is None:
        end_date = today

    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="结束时间不能早于开始时间",
        )
    if (end_date - start_date).days > 90:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="时间范围不能超过90天，请缩小范围",
        )

    # 不分页获取全量数据
    data = await stats_service.get_report_data(
        report_type, start_date, end_date, 1, 9999, db
    )

    # 用 openpyxl 生成 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = report_type

    headers = _REPORT_HEADERS.get(report_type, ["日期"])
    fields = _REPORT_FIELDS.get(report_type, ["date"])

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, item in enumerate(data.get("list", []), 2):
        for col_idx, field in enumerate(fields, 1):
            val = item.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")

    # 自动列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today_str = today.strftime("%Y%m%d")
    filename = f"report_{report_type}_{today_str}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/stats/liblib",
    dependencies=[require_role(*_LIBLIB_READ_ROLES)],
)
async def get_liblib_stats(
    days: int = Query(7, ge=1, le=30, description="统计近 N 天"),
    admin_user: AdminUser = Depends(get_current_admin),
):
    """LiblibAI 调用统计看板（STEP-036 · PRD 10.7#1）：读 Redis liblib_stats:{YYYYMMDD} HSET。

    tech_ops 可只读本看板；返回近 days 天日汇总（total/success/failed/points_used）+ 合计。
    """
    from backend.redis_client import get_redis

    today = datetime.date.today()
    daily = []
    sum_total = sum_success = sum_failed = sum_points = 0
    try:
        r = await get_redis()
        for i in range(days - 1, -1, -1):
            d = today - datetime.timedelta(days=i)
            key = f"liblib_stats:{d.strftime('%Y%m%d')}"
            h = await r.hgetall(key) or {}

            def _int(v):
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return 0

            total = _int(h.get("total"))
            success = _int(h.get("success"))
            failed = _int(h.get("failed"))
            points = _int(h.get("points_used"))
            sum_total += total
            sum_success += success
            sum_failed += failed
            sum_points += points
            daily.append({
                "date": d.strftime("%Y-%m-%d"),
                "total": total,
                "success": success,
                "failed": failed,
                "points_used": points,
            })
    except Exception as e:
        logger.error("[看板] 读取 liblib_stats 失败: %s", e)
        return ApiResponse.ok(data={"days": days, "daily": [], "summary": {}, "redis_error": True})

    return ApiResponse.ok(data={
        "days": days,
        "daily": daily,
        "summary": {
            "total": sum_total,
            "success": sum_success,
            "failed": sum_failed,
            "points_used": sum_points,
        },
    })
