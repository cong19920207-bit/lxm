# -*- coding: utf-8 -*-
# 系统监控：系统状态、第三方服务状态

import datetime
import io
import json
import logging
import os
import time
from typing import Optional

import psutil
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.models.admin_config import AdminConfig
from backend.models.admin_user import AdminUser
from backend.redis_client import get_redis
from backend.schemas.common import ApiResponse
from backend.utils.admin_auth import (
    deny_observer_export,
    get_current_admin,
    log_operation,
    require_role,
)
from backend.utils.credential_redaction import REDACTED, redact_credentials

logger = logging.getLogger(__name__)

router = APIRouter()

_SYSTEM_READ_ROLES = ("super_admin", "tech_ops", "observer")
_SYSTEM_WRITE_ROLES = ("super_admin", "tech_ops")
_SYSTEM_EXPORT_ROLES = ("super_admin", "tech_ops")


async def _get_cached(redis_client, cache_key: str):
    """从 Redis 读取缓存的 JSON 数据，不存在返回 None"""
    raw = await redis_client.get(cache_key)
    if raw:
        try:
            return json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            pass
    return None


async def _set_cached(redis_client, cache_key: str, data: dict, ttl: int):
    """将数据序列化为 JSON 写入 Redis 缓存"""
    await redis_client.set(cache_key, json.dumps(data, ensure_ascii=False), ex=ttl)


async def _observer_credential_status(data: dict, db: AsyncSession) -> dict:
    """为 observer 复制第三方状态并仅附加凭据是否已配置。"""
    service_names = {
        "LLM（豆包）": "doubao",
        "Embedding（阿里云）": "embedding",
        "DashVector（向量检索）": "dashvector",
    }
    result = {**data, "services": [dict(item) for item in data.get("services", [])]}
    for item in result["services"]:
        service_name = service_names.get(item.get("name"))
        if service_name is None:
            continue
        _, config = await _get_active_third_party_record(db, service_name)
        item["credential_configured"] = any(
            bool(config.get(field)) for field in _SENSITIVE_FIELDS
        )
    return result


@router.get(
    "/system/status",
    dependencies=[require_role(*_SYSTEM_READ_ROLES)],
)
async def get_system_status():
    """
    系统状态监控：CPU/内存/磁盘 + Redis 命中率 + 告警列表。
    结果缓存10秒（key=cache:system_status）。
    """
    r = await get_redis()
    cache_key = "cache:system_status"

    cached = await _get_cached(r, cache_key)
    if cached:
        return ApiResponse.ok(data=cached)

    # ---- 采集系统指标 ----
    cpu_percent = psutil.cpu_percent(interval=0.5)
    mem = psutil.virtual_memory()
    disk = psutil.disk_usage("/")

    # ---- Redis INFO ----
    redis_hit_rate = 0.0
    redis_used_memory = "0M"
    redis_connected_clients = 0
    try:
        info = await r.info(section="stats")
        info_mem = await r.info(section="memory")
        info_clients = await r.info(section="clients")

        hits = int(info.get("keyspace_hits", 0))
        misses = int(info.get("keyspace_misses", 0))
        if hits + misses > 0:
            redis_hit_rate = round(hits / (hits + misses) * 100, 2)

        used_bytes = info_mem.get("used_memory", 0)
        redis_used_memory = f"{round(int(used_bytes) / 1024 / 1024, 1)}M"
        redis_connected_clients = int(info_clients.get("connected_clients", 0))
    except Exception as e:
        logger.warning("获取 Redis INFO 失败: %s", e)

    # ---- 生成告警 ----
    alerts = []
    if cpu_percent > 90:
        alerts.append({"level": "critical", "message": f"CPU 使用率过高: {cpu_percent}%"})
    elif cpu_percent > 70:
        alerts.append({"level": "warning", "message": f"CPU 使用率偏高: {cpu_percent}%"})

    if mem.percent > 90:
        alerts.append({"level": "critical", "message": f"内存使用率过高: {mem.percent}%"})
    elif mem.percent > 70:
        alerts.append({"level": "warning", "message": f"内存使用率偏高: {mem.percent}%"})

    if disk.percent > 90:
        alerts.append({"level": "critical", "message": f"磁盘使用率过高: {disk.percent}%"})
    elif disk.percent > 80:
        alerts.append({"level": "warning", "message": f"磁盘使用率偏高: {disk.percent}%"})

    if redis_hit_rate < 50 and (hits + misses) > 100:
        alerts.append({"level": "warning", "message": f"Redis 缓存命中率偏低: {redis_hit_rate}%"})

    data = {
        "cpu": {
            "percent": cpu_percent,
            "cores": psutil.cpu_count(),
        },
        "memory": {
            "percent": mem.percent,
            "total_gb": round(mem.total / 1024 / 1024 / 1024, 1),
            "used_gb": round(mem.used / 1024 / 1024 / 1024, 1),
        },
        "disk": {
            "percent": disk.percent,
            "total_gb": round(disk.total / 1024 / 1024 / 1024, 1),
            "used_gb": round(disk.used / 1024 / 1024 / 1024, 1),
        },
        "redis": {
            "hit_rate": redis_hit_rate,
            "used_memory": redis_used_memory,
            "connected_clients": redis_connected_clients,
        },
        "alerts": alerts,
    }

    await _set_cached(r, cache_key, data, ttl=10)
    return ApiResponse.ok(data=data)


@router.get(
    "/third-party/status",
    dependencies=[require_role(*_SYSTEM_READ_ROLES)],
)
async def get_third_party_status(
    admin_user: AdminUser = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    第三方服务状态：LLM / Embedding / DashVector / 内容安全。
    从 Redis 统计 key 汇总，结果缓存60秒（key=cache:third_party_status）。
    """
    r = await get_redis()
    cache_key = "cache:third_party_status"

    cached = await _get_cached(r, cache_key)
    if cached:
        if admin_user.role == "observer":
            cached = await _observer_credential_status(cached, db)
        return ApiResponse.ok(data=cached)

    today_str = datetime.date.today().strftime("%Y%m%d")

    # ---- LLM（豆包）----
    llm_total = 0
    llm_success = 0
    llm_avg_ms = 0.0
    llm_status = "normal"
    try:
        total_raw = await r.hget(f"llm_stats:{today_str}", "total")
        success_raw = await r.hget(f"llm_stats:{today_str}", "success")
        llm_total = int(total_raw) if total_raw else 0
        llm_success = int(success_raw) if success_raw else 0

        response_times = await r.lrange("llm_response_times", 0, 99)
        if response_times:
            times = [float(t) for t in response_times]
            llm_avg_ms = round(sum(times) / len(times), 1)

        if llm_total > 0:
            success_rate = llm_success / llm_total * 100
            if success_rate < 90:
                llm_status = "error"
            elif success_rate < 95:
                llm_status = "warning"
        elif llm_total == 0:
            llm_status = "unknown"
    except Exception as e:
        logger.warning("读取 LLM 统计失败: %s", e)
        llm_status = "error"

    # ---- Embedding（阿里云）----
    emb_total = 0
    emb_success = 0
    emb_avg_ms = 0.0
    emb_status = "normal"
    try:
        emb_total_raw = await r.hget(f"embedding_stats:{today_str}", "total")
        emb_success_raw = await r.hget(f"embedding_stats:{today_str}", "success")
        emb_total = int(emb_total_raw) if emb_total_raw else 0
        emb_success = int(emb_success_raw) if emb_success_raw else 0

        emb_times_raw = await r.lrange("embedding_response_times", 0, 99)
        if emb_times_raw:
            emb_times = [float(t) for t in emb_times_raw]
            emb_avg_ms = round(sum(emb_times) / len(emb_times), 1)

        if emb_total > 0:
            rate = emb_success / emb_total * 100
            if rate < 90:
                emb_status = "error"
            elif rate < 95:
                emb_status = "warning"
        elif emb_total == 0:
            emb_status = "unknown"
    except Exception as e:
        logger.warning("读取 Embedding 统计失败: %s", e)
        emb_status = "error"

    # ---- DashVector（向量检索）----
    vec_total = 0
    vec_success = 0
    vec_avg_ms = 0.0
    vec_status = "normal"
    try:
        vec_total_raw = await r.hget(f"vector_stats:{today_str}", "total")
        vec_success_raw = await r.hget(f"vector_stats:{today_str}", "success")
        vec_total = int(vec_total_raw) if vec_total_raw else 0
        vec_success = int(vec_success_raw) if vec_success_raw else 0

        vec_times_raw = await r.lrange("vector_response_times", 0, 99)
        if vec_times_raw:
            vec_times = [float(t) for t in vec_times_raw]
            vec_avg_ms = round(sum(vec_times) / len(vec_times), 1)

        if vec_total > 0:
            rate = vec_success / vec_total * 100
            if rate < 90:
                vec_status = "error"
            elif rate < 95:
                vec_status = "warning"
        elif vec_total == 0:
            vec_status = "unknown"
    except Exception as e:
        logger.warning("读取 DashVector 统计失败: %s", e)
        vec_status = "error"

    # ---- 内容安全 ----
    block_count = 0
    safety_status = "normal"
    try:
        block_raw = await r.get(f"content_block_count:{today_str}")
        block_count = int(block_raw) if block_raw else 0
        if block_count > 50:
            safety_status = "warning"
    except Exception as e:
        logger.warning("读取内容安全统计失败: %s", e)
        safety_status = "error"

    data = {
        "services": [
            {
                "name": "LLM（豆包）",
                "status": llm_status,
                "today_calls": llm_total,
                "today_success": llm_success,
                "success_rate": round(llm_success / llm_total * 100, 1) if llm_total > 0 else 0,
                "avg_response_ms": llm_avg_ms,
            },
            {
                "name": "Embedding（阿里云）",
                "status": emb_status,
                "today_calls": emb_total,
                "today_success": emb_success,
                "success_rate": round(emb_success / emb_total * 100, 1) if emb_total > 0 else 0,
                "avg_response_ms": emb_avg_ms,
            },
            {
                "name": "DashVector（向量检索）",
                "status": vec_status,
                "today_calls": vec_total,
                "today_success": vec_success,
                "success_rate": round(vec_success / vec_total * 100, 1) if vec_total > 0 else 0,
                "avg_response_ms": vec_avg_ms,
            },
            {
                "name": "内容安全",
                "status": safety_status,
                "today_blocked": block_count,
            },
        ],
    }

    await _set_cached(r, cache_key, data, ttl=60)
    if admin_user.role == "observer":
        data = await _observer_credential_status(data, db)
    return ApiResponse.ok(data=data)


# ============ 第三方服务配置与测试 ============

_VALID_SERVICES = {"doubao", "embedding", "dashvector", "content_safety"}

_SERVICE_CONFIG_KEY_MAP = {
    "doubao": "third_party:doubao",
    "embedding": "third_party:embedding",
    "dashvector": "third_party:dashvector",
    "content_safety": "third_party:content_safety",
}

_SENSITIVE_FIELDS = {"api_key", "secret_key", "api_secret", "password", "token"}


def _mask_sensitive(config: dict) -> dict:
    """将配置中的敏感字段值替换为掩码，不暴露明文 API Key"""
    masked = {}
    for k, v in config.items():
        if k in _SENSITIVE_FIELDS and isinstance(v, str) and len(v) > 4:
            masked[k] = v[:3] + "*" * (len(v) - 6) + v[-3:] if len(v) > 6 else "***"
        else:
            masked[k] = v
    return masked


async def _get_active_third_party_record(
    db: AsyncSession,
    service_name: str,
) -> tuple[AdminConfig | None, dict]:
    """
    读取指定第三方服务当前已发布配置行与解析后的 JSON。
    无生效行或解析失败时 config 为空 dict，current 可能为 None 或无效行。
    """
    if service_name not in _VALID_SERVICES:
        return None, {}
    config_key = _SERVICE_CONFIG_KEY_MAP[service_name]
    stmt = select(AdminConfig).where(
        AdminConfig.config_key == config_key,
        AdminConfig.is_active == True,   # noqa: E712
        AdminConfig.is_draft == False,   # noqa: E712
    )
    result = await db.execute(stmt)
    current = result.scalars().first()
    if not current or not current.config_value:
        return current, {}
    try:
        cfg = json.loads(current.config_value)
        if not isinstance(cfg, dict):
            return current, {}
        return current, cfg
    except (json.JSONDecodeError, TypeError):
        return current, {}


@router.put(
    "/third-party/{service_name}/config",
    dependencies=[require_role(*_SYSTEM_WRITE_ROLES)],
)
async def update_third_party_config(
    service_name: str,
    body: dict,
    request: Request,
    admin_user: AdminUser = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    更新第三方服务配置。
    先用新配置测试连接，通过后合并写入 admin_config 并记录操作日志。
    响应中不返回明文 API Key。
    """
    if service_name not in _VALID_SERVICES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的服务名，可选值: {', '.join(sorted(_VALID_SERVICES))}",
        )

    if not body:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="请求体不能为空",
        )

    config_key = _SERVICE_CONFIG_KEY_MAP[service_name]
    current, existing_config = await _get_active_third_party_record(db, service_name)

    before_value = json.dumps(existing_config, ensure_ascii=False) if existing_config else None

    # 合并新配置
    merged_config = {**existing_config, **body}

    # 先测试连接
    test_result = await _do_test_connection(service_name, merged_config)
    if not test_result["connected"]:
        return ApiResponse.fail(
            code=5001,
            message=f"连接测试失败，配置未保存: {test_result.get('error', '未知错误')}",
        )

    # 测试通过，保存到 admin_config
    new_value = json.dumps(merged_config, ensure_ascii=False)
    now = datetime.datetime.utcnow()

    if current:
        current.is_active = False

    new_version = (current.version + 1) if current else 1
    new_config = AdminConfig(
        config_key=config_key,
        config_value=new_value,
        version=new_version,
        is_draft=False,
        is_active=True,
        updated_by=admin_user.username,
        updated_at=now,
    )
    db.add(new_config)
    await db.flush()

    # 更新 Redis 缓存
    r = await get_redis()
    await r.setex(f"active_config:{config_key}", 3600, new_value)

    # 写操作日志
    await log_operation(
        db=db,
        admin_user=admin_user,
        module="third_party",
        action="update_config",
        target_description=f"更新第三方服务配置: {service_name}",
        before_value=before_value,
        after_value=json.dumps(_mask_sensitive(merged_config), ensure_ascii=False),
        request=request,
    )

    return ApiResponse.ok(data={
        "service_name": service_name,
        "config": _mask_sensitive(merged_config),
        "version": new_version,
        "test_latency_ms": test_result.get("latency_ms", 0),
    })


@router.post(
    "/third-party/{service_name}/test-connection",
    dependencies=[require_role(*_SYSTEM_WRITE_ROLES)],
)
async def test_third_party_connection(
    service_name: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """
    测试第三方服务连通性：与 PUT config 相同，将「已发布配置」与请求体 JSON 合并后再测。
    请求体可选；无 body 或 {} 时仅用已发布配置 + 各 _test_* 内对环境变量的回退。
    """
    if service_name not in _VALID_SERVICES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的服务名，可选值: {', '.join(sorted(_VALID_SERVICES))}",
        )

    patch: dict = {}
    raw = await request.body()
    if raw:
        try:
            obj = json.loads(raw.decode("utf-8"))
            if isinstance(obj, dict):
                patch = obj
        except (json.JSONDecodeError, UnicodeDecodeError, TypeError):
            patch = {}

    _, existing_config = await _get_active_third_party_record(db, service_name)
    merged_config = {**existing_config, **patch}
    result = await _do_test_connection(service_name, merged_config)
    return ApiResponse.ok(data=result)


async def _do_test_connection(service_name: str, config_override: dict = None) -> dict:
    """
    执行实际的连接测试。
    config_override 不为空时用传入的配置创建临时客户端，否则用全局默认配置。
    """
    start_ms = time.monotonic()
    result = {"connected": False, "latency_ms": 0, "error": ""}

    try:
        if service_name == "doubao":
            await _test_doubao(config_override)
        elif service_name == "embedding":
            await _test_embedding(config_override)
        elif service_name == "dashvector":
            await _test_dashvector(config_override)
        elif service_name == "content_safety":
            await _test_content_safety()

        result["connected"] = True
    except Exception as e:
        result["error"] = str(e)
        logger.warning("第三方服务连接测试失败 [%s]: %s", service_name, e)

    result["latency_ms"] = int((time.monotonic() - start_ms) * 1000)
    return result


async def _test_doubao(config: dict = None):
    """豆包 LLM 最小测试：发送一句话，验证能收到回复"""
    import httpx
    from backend.config import get_volc_api_key, get_volc_endpoint, get_volc_model

    api_key = (config or {}).get("api_key") or get_volc_api_key()
    endpoint = (config or {}).get("endpoint") or get_volc_endpoint()
    model = (config or {}).get("model") or get_volc_model()

    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.post(
            f"{endpoint}/chat/completions",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            json={
                "model": model,
                "messages": [{"role": "user", "content": "你好"}],
                "stream": False,
                "max_tokens": 10,
            },
        )
        resp.raise_for_status()
        data = resp.json()
        choices = data.get("choices", [])
        if not choices:
            raise RuntimeError("LLM 返回无 choices")


async def _test_embedding(config: dict = None):
    """Embedding 最小测试：向量化一个短文本"""
    import httpx
    from backend.config import get_aliyun_api_key, get_embedding_endpoint, get_embedding_model

    api_key = (config or {}).get("api_key") or get_aliyun_api_key()
    endpoint = (config or {}).get("endpoint") or get_embedding_endpoint()
    model = (config or {}).get("model") or get_embedding_model()

    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.post(
            endpoint,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            json={
                "model": model,
                "input": {"texts": ["连接测试"]},
                "parameters": {"text_type": "query"},
            },
        )
        resp.raise_for_status()
        data = resp.json()
        embeddings = data.get("output", {}).get("embeddings", [])
        if not embeddings:
            raise RuntimeError("Embedding 返回为空")


async def _test_dashvector(config: dict = None):
    """DashVector 最小测试：调用 collection 描述接口验证连通性"""
    import httpx
    from backend.config import (
        get_dashvector_api_key, get_dashvector_collection, get_dashvector_endpoint,
    )

    api_key = (config or {}).get("api_key") or get_dashvector_api_key()
    endpoint = (config or {}).get("endpoint") or get_dashvector_endpoint()
    collection = (config or {}).get("collection") or get_dashvector_collection()

    async with httpx.AsyncClient(timeout=10.0) as client:
        resp = await client.get(
            f"{endpoint}/v1/collections/{collection}",
            headers={
                "Content-Type": "application/json",
                "dashvector-auth-token": api_key,
            },
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") != 0:
            raise RuntimeError(f"DashVector 返回错误: {data.get('message', '未知')}")


async def _test_content_safety():
    """内容安全最小测试：验证 Redis 中 banned_keywords 可读取"""
    r = await get_redis()
    await r.get("banned_keywords")


# ============ 系统日志查询与导出 ============

_LOG_DIR = "logs"

_LOG_TYPE_FILE_MAP = {
    "system": "system.log",
    "error": "error.log",
}

_VALID_LEVELS = {"DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"}


def _redact_system_log_entry(entry: dict) -> dict:
    try:
        redacted = redact_credentials(entry)
        if isinstance(redacted, dict):
            return redacted
    except Exception:
        logger.exception("系统日志读取脱敏失败，消息已按失败关闭处理")
    return {**entry, "message": REDACTED}


def _parse_log_line(line: str) -> dict | None:
    """
    解析单行日志，格式: 2026-04-01 12:00:00 | INFO | module | message
    返回 {"time": str, "level": str, "module": str, "message": str} 或 None
    """
    parts = line.split(" | ", 3)
    if len(parts) < 4:
        return None
    time_str = parts[0].strip()
    level = parts[1].strip()
    module = parts[2].strip()
    message = parts[3].strip()
    if level not in _VALID_LEVELS:
        return None
    return {"time": time_str, "level": level, "module": module, "message": message}


def _collect_log_files(
    base_filename: str,
    start_date: datetime.date,
    end_date: datetime.date,
) -> list[str]:
    """
    收集指定日期范围内的日志文件路径列表。
    TimedRotatingFileHandler(when="midnight") 生成的轮转文件命名:
      system.log          （当天）
      system.log.2026-03-31（历史）
    返回按日期从新到旧排序的文件路径列表。
    """
    today = datetime.date.today()
    files = []

    if start_date <= today <= end_date:
        current_file = os.path.join(_LOG_DIR, base_filename)
        if os.path.isfile(current_file):
            files.append((today, current_file))

    d = end_date
    while d >= start_date:
        if d != today:
            suffix = d.strftime("%Y-%m-%d")
            rotated = os.path.join(_LOG_DIR, f"{base_filename}.{suffix}")
            if os.path.isfile(rotated):
                files.append((d, rotated))
        d -= datetime.timedelta(days=1)

    files.sort(key=lambda x: x[0], reverse=True)
    return [f[1] for f in files]


def _read_and_filter_logs(
    file_paths: list[str],
    level: str | None,
    start_date: datetime.date,
    end_date: datetime.date,
) -> list[dict]:
    """读取多个日志文件并按条件过滤，按 time 降序（最新在前）返回"""
    all_entries = []
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    for fp in file_paths:
        try:
            with open(fp, "r", encoding="utf-8", errors="replace") as f:
                for raw_line in f:
                    raw_line = raw_line.rstrip("\n\r")
                    if not raw_line:
                        continue
                    entry = _parse_log_line(raw_line)
                    if entry is None:
                        continue
                    if level and entry["level"] != level:
                        continue
                    date_part = entry["time"][:10]
                    if date_part < start_str or date_part > end_str:
                        continue
                    all_entries.append(entry)
        except Exception as e:
            logger.warning("读取日志文件失败 [%s]: %s", fp, e)

    # 跨多个轮转文件时不能依赖拼接顺序 + reverse，须按时间戳全局降序
    all_entries.sort(key=lambda e: e["time"], reverse=True)
    return all_entries


@router.get(
    "/system/logs",
    dependencies=[require_role(*_SYSTEM_READ_ROLES)],
)
async def get_system_logs(
    log_type: str = Query("system", description="日志类型: system / error"),
    level: Optional[str] = Query(None, description="日志级别: DEBUG/INFO/WARNING/ERROR/CRITICAL"),
    start_date: Optional[datetime.date] = Query(None, description="开始日期"),
    end_date: Optional[datetime.date] = Query(None, description="结束日期"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(50, ge=1, le=200, description="每页条数"),
):
    """
    按条件查询系统日志文件，按行解析后倒序分页返回。
    """
    if log_type not in _LOG_TYPE_FILE_MAP:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"log_type 参数无效，可选值: {', '.join(_LOG_TYPE_FILE_MAP.keys())}",
        )

    if level and level.upper() not in _VALID_LEVELS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"level 参数无效，可选值: {', '.join(sorted(_VALID_LEVELS))}",
        )
    if level:
        level = level.upper()

    today = datetime.date.today()
    if start_date is None:
        start_date = today - datetime.timedelta(days=7)
    if end_date is None:
        end_date = today

    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="结束日期不能早于开始日期",
        )
    if (end_date - start_date).days > 30:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="查询范围不能超过30天",
        )

    base_filename = _LOG_TYPE_FILE_MAP[log_type]
    file_paths = _collect_log_files(base_filename, start_date, end_date)

    if not file_paths:
        return ApiResponse.ok(data={
            "total": 0, "page": page, "page_size": page_size, "list": [],
        })

    entries = _read_and_filter_logs(file_paths, level, start_date, end_date)

    total = len(entries)
    offset = (page - 1) * page_size
    paged = [
        _redact_system_log_entry(entry)
        for entry in entries[offset: offset + page_size]
    ]

    return ApiResponse.ok(data={
        "total": total,
        "page": page,
        "page_size": page_size,
        "list": paged,
    })


@router.post(
    "/system/logs/export",
    dependencies=[
        Depends(deny_observer_export),
        require_role(*_SYSTEM_EXPORT_ROLES),
    ],
)
async def export_system_logs(
    log_type: str = Query("system", description="日志类型: system / error"),
    level: Optional[str] = Query(None, description="日志级别过滤"),
    start_date: Optional[datetime.date] = Query(None, description="开始日期"),
    end_date: Optional[datetime.date] = Query(None, description="结束日期"),
):
    """
    导出系统日志为 Excel 文件。同 GET /system/logs 条件但不分页，单次限7天。
    """
    if log_type not in _LOG_TYPE_FILE_MAP:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"log_type 参数无效，可选值: {', '.join(_LOG_TYPE_FILE_MAP.keys())}",
        )

    if level and level.upper() not in _VALID_LEVELS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"level 参数无效，可选值: {', '.join(sorted(_VALID_LEVELS))}",
        )
    if level:
        level = level.upper()

    today = datetime.date.today()
    if start_date is None:
        start_date = today - datetime.timedelta(days=7)
    if end_date is None:
        end_date = today

    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="结束日期不能早于开始日期",
        )
    if (end_date - start_date).days > 7:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="导出范围不能超过7天，请缩小日期范围",
        )

    base_filename = _LOG_TYPE_FILE_MAP[log_type]
    file_paths = _collect_log_files(base_filename, start_date, end_date)
    entries = [
        _redact_system_log_entry(entry)
        for entry in _read_and_filter_logs(file_paths, level, start_date, end_date)
    ]

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"{log_type}_logs"

    headers = ["时间", "级别", "模块", "消息"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    fields = ["time", "level", "module", "message"]
    for row_idx, entry in enumerate(entries, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=entry.get(field, ""))

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{log_type}_logs_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# TODO: 后续接口
