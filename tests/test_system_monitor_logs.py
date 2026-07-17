# -*- coding: utf-8 -*-
# 管理后台系统日志：文件解析与多文件时间排序（不触库）

import datetime
import io
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.routers.admin import system_monitor
from backend.routers.admin.system_monitor import (
    _parse_log_line,
    _read_and_filter_logs,
    export_system_logs,
    get_system_logs,
)
from backend.utils.credential_redaction import REDACTED


def test_parse_log_line_valid():
    line = "2026-05-10 12:00:00 | INFO | some.module | hello | world"
    out = _parse_log_line(line)
    assert out is not None
    assert out["time"] == "2026-05-10 12:00:00"
    assert out["level"] == "INFO"
    assert out["module"] == "some.module"
    assert out["message"] == "hello | world"


def test_parse_log_line_invalid_short():
    assert _parse_log_line("no pipes") is None


@pytest.fixture()
def two_day_log_files():
    """模拟当日文件 + 昨日轮转文件，条目不按时间交错写入。"""
    with tempfile.TemporaryDirectory() as tmp:
        today_p = Path(tmp) / "system.log"
        yday_p = Path(tmp) / "system.log.2026-05-09"
        # 当日文件：较早的一条在前（文件中升序）
        today_p.write_text(
            "2026-05-10 10:00:00 | INFO | a | early today\n"
            "2026-05-10 18:00:00 | INFO | a | late today\n",
            encoding="utf-8",
        )
        # 昨日文件：午夜一条
        yday_p.write_text(
            "2026-05-09 23:59:00 | INFO | a | yesterday last\n",
            encoding="utf-8",
        )
        # 与 _collect_log_files 一致：先今日路径再昨日路径
        yield [str(today_p), str(yday_p)]


def test_read_and_filter_logs_sorts_newest_first(two_day_log_files):
    start = datetime.date(2026, 5, 9)
    end = datetime.date(2026, 5, 10)
    rows = _read_and_filter_logs(two_day_log_files, None, start, end)
    times = [r["time"] for r in rows]
    assert times == sorted(times, reverse=True)
    assert rows[0]["message"] == "late today"
    assert rows[-1]["message"] == "yesterday last"


def test_read_and_filter_logs_level_filter(two_day_log_files):
    start = datetime.date(2026, 5, 9)
    end = datetime.date(2026, 5, 10)
    rows = _read_and_filter_logs(two_day_log_files, "WARNING", start, end)
    assert rows == []


@pytest.mark.asyncio
async def test_system_log_list_and_export_share_credential_redaction(tmp_path, monkeypatch):
    day = datetime.date(2026, 7, 16)
    log_file = tmp_path / "system.log.2026-07-16"
    log_file.write_text(
        "2026-07-16 10:00:00 | INFO | auth | api_key=system-api-key prompt=keep-prompt\n"
        "2026-07-16 11:00:00 | WARNING | client | Authorization: Bearer system-bearer-token\n"
        "2026-07-16 12:00:00 | INFO | config | max_tokens=128 description=keep-description\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        system_monitor,
        "_collect_log_files",
        lambda _base, _start, _end: [str(log_file)],
    )

    listed = await get_system_logs(
        log_type="system",
        level=None,
        start_date=day,
        end_date=day,
        page=1,
        page_size=50,
    )
    exported = await export_system_logs(
        log_type="system",
        level=None,
        start_date=day,
        end_date=day,
    )

    list_text = "\n".join(entry["message"] for entry in listed.data["list"])
    export_bytes = b"".join([chunk async for chunk in exported.body_iterator])
    workbook = load_workbook(io.BytesIO(export_bytes), read_only=True)
    export_text = "\n".join(
        str(value)
        for row in workbook.active.iter_rows(values_only=True)
        for value in row
        if value is not None
    )

    for rendered in (list_text, export_text):
        assert REDACTED in rendered
        assert "system-api-key" not in rendered
        assert "system-bearer-token" not in rendered
        assert "keep-prompt" in rendered
        assert "max_tokens=128" in rendered
        assert "keep-description" in rendered


@pytest.mark.asyncio
async def test_system_log_redaction_failure_closes_message_only(tmp_path, monkeypatch):
    day = datetime.date(2026, 7, 16)
    log_file = tmp_path / "system.log.2026-07-16"
    log_file.write_text(
        "2026-07-16 10:00:00 | ERROR | auth | password=must-not-leak\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        system_monitor,
        "_collect_log_files",
        lambda _base, _start, _end: [str(log_file)],
    )
    monkeypatch.setattr(
        system_monitor,
        "redact_credentials",
        lambda _value: (_ for _ in ()).throw(RuntimeError("simulated failure")),
        raising=False,
    )

    listed = await get_system_logs(
        log_type="system",
        level=None,
        start_date=day,
        end_date=day,
        page=1,
        page_size=50,
    )

    assert listed.data["list"] == [
        {
            "time": "2026-07-16 10:00:00",
            "level": "ERROR",
            "module": "auth",
            "message": REDACTED,
        }
    ]
