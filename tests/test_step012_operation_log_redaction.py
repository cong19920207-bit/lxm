# -*- coding: utf-8 -*-

import io
import json

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from backend.database import Base
from backend.models.admin_operation_log import AdminOperationLog
from backend.models.admin_user import AdminUser
from backend.routers.admin.operation_logs import (
    export_operation_logs,
    get_operation_log_detail,
    list_operation_logs,
)
from backend.utils.admin_auth import log_operation
from backend.utils.credential_redaction import REDACTED


engine = create_async_engine("sqlite+aiosqlite:///:memory:")
session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


@pytest_asyncio.fixture(autouse=True)
async def database():
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    yield
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.drop_all)


async def _create_admin(session: AsyncSession) -> AdminUser:
    admin = AdminUser(
        username="step012-admin",
        password_hash="unused",
        role="super_admin",
        token_version=0,
    )
    session.add(admin)
    await session.flush()
    return admin


@pytest.mark.asyncio
async def test_log_operation_redacts_before_database_write():
    async with session_factory() as session:
        admin = await _create_admin(session)

        await log_operation(
            db=session,
            admin_user=admin,
            module="security",
            action="edit",
            target_description="rotate api_key=target-secret version=v2",
            before_value=json.dumps(
                {"api_key": "before-secret", "prompt": "keep prompt"},
                ensure_ascii=False,
            ),
            after_value=json.dumps(
                {"nested": [{"refresh_token": "after-secret"}], "max_tokens": 12},
                ensure_ascii=False,
            ),
        )

        stored = (await session.execute(select(AdminOperationLog))).scalar_one()
        assert stored.target_description == f"rotate api_key={REDACTED} version=v2"
        assert json.loads(stored.before_value) == {
            "api_key": REDACTED,
            "prompt": "keep prompt",
        }
        assert json.loads(stored.after_value) == {
            "nested": [{"refresh_token": REDACTED}],
            "max_tokens": 12,
        }


@pytest.mark.asyncio
async def test_historical_plaintext_is_redacted_in_list_detail_and_export_only():
    async with session_factory() as session:
        admin = await _create_admin(session)
        historical = AdminOperationLog(
            admin_user_id=admin.id,
            admin_username=admin.username,
            module="security",
            action="edit",
            target_description="Authorization: Bearer historical-target-token",
            before_value=json.dumps(
                {"password": "historical-password", "description": "keep description"},
                ensure_ascii=False,
            ),
            after_value="private_key='historical-private-key'; version=3",
            ip_address="127.0.0.1",
        )
        session.add(historical)
        await session.flush()
        log_id = historical.id

        listed = await list_operation_logs(
            admin_username=None,
            module=None,
            action=None,
            start_date=None,
            end_date=None,
            page=1,
            page_size=20,
            db=session,
            admin_user=admin,
        )
        detailed = await get_operation_log_detail(log_id, db=session, admin_user=admin)
        exported = await export_operation_logs(
            admin_username=None,
            module=None,
            action=None,
            start_date=None,
            end_date=None,
            db=session,
            admin_user=admin,
        )

        assert listed.data["list"][0]["target_description"] == f"Authorization: {REDACTED}"
        assert json.loads(detailed.data["before_value"]) == {
            "password": REDACTED,
            "description": "keep description",
        }
        assert detailed.data["after_value"] == f"private_key='{REDACTED}'; version=3"

        export_bytes = b"".join([chunk async for chunk in exported.body_iterator])
        workbook = load_workbook(io.BytesIO(export_bytes), read_only=True)
        values = [value for row in workbook.active.iter_rows(values_only=True) for value in row]
        rendered = "\n".join(str(value) for value in values if value is not None)
        assert REDACTED in rendered
        for plaintext in (
            "historical-target-token",
            "historical-password",
            "historical-private-key",
        ):
            assert plaintext not in rendered
        assert "keep description" in rendered
        assert "version=3" in rendered

        await session.refresh(historical)
        assert "historical-target-token" in historical.target_description
        assert "historical-password" in historical.before_value
        assert "historical-private-key" in historical.after_value


@pytest.mark.asyncio
async def test_redaction_exception_does_not_abort_audit_write(monkeypatch):
    from backend.utils import admin_auth

    def fail_redaction(_value):
        raise RuntimeError("simulated redaction failure")

    monkeypatch.setattr(admin_auth, "redact_credentials", fail_redaction, raising=False)

    async with session_factory() as session:
        admin = await _create_admin(session)
        await log_operation(
            db=session,
            admin_user=admin,
            module="security",
            action="edit",
            target_description="api_key=must-not-leak",
            before_value="password=must-not-leak",
            after_value="Bearer must-not-leak-token",
        )

        stored = (await session.execute(select(AdminOperationLog))).scalar_one()
        assert stored.target_description == REDACTED
        assert stored.before_value == REDACTED
        assert stored.after_value == REDACTED
